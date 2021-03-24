import frappe
from frappe.utils.pdf import get_pdf
import openpyxl
from openpyxl.styles import Alignment
import datetime


@frappe.whitelist()
def get_report_data(from_date, from_shift, to_date, to_shift):
    print("\n\n\n", from_date, from_shift, to_date, to_shift)
    print("\n\n\n\n\n\n\n\n\nMethod is called\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n")
    get_data = frappe.db.sql("""SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('{}','Evening') AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date>'{}' AND  transaction_date<'{}' AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('Morning','{}') AND docstatus=1
                        """.format(from_date, from_shift, from_date, to_date, to_date, to_shift), as_dict=1)
    # print(get_data,"\n\n\n\n\n\n")
    data = []
    data.extend(get_data)
    print("\n\n\n\n\n", data, "\n\n\n\n\n\n\n\n")
    context = {
        "records": data
    }
    template = frappe.render_template(
        "agency_management/public/js/page_templates/quotation_list.html", context)
#     print(template,"\n\n\n")
    return template


@frappe.whitelist()
def create_pdf(from_date, from_shift, to_date, to_shift):
    get_data = frappe.db.sql("""SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('{}','Evening') AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date>'{}' AND  transaction_date<'{}' AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('Morning','{}') AND docstatus=1
                        """.format(from_date, from_shift, from_date, to_date, to_date, to_shift), as_dict=1)
    data = []
    data.extend(get_data)
    context = {
        "records": data
    }
    template = frappe.render_template(
        "agency_management/public/js/page_templates/page_print_format.html", context)
    pdf = get_pdf(template)
    with open('/home/ruturaj/Downloads/report.pdf', 'wb') as file:
        file.write(pdf)
    return pdf


@frappe.whitelist()
def export_excel(from_date, from_shift, to_date, to_shift):
    # frappe.throw("Export to Excel Method is called")
    get_data = frappe.db.sql("""SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('{}','Evening') AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date>'{}' AND  transaction_date<'{}' AND docstatus=1
                            UNION SELECT quotation_to, customer_name, transaction_date, shift, valid_till, contact_mobile, total, total_taxes_and_charges, grand_total FROM `tabQuotation`
                                    WHERE transaction_date='{}' AND shift IN('Morning','{}') AND docstatus=1
                        """.format(from_date, from_shift, from_date, to_date, to_date, to_shift), as_dict=1)
    datas = []
    datas.extend(get_data)
    print("\n\n\n\n\n\nData is fetched\n\n\n\n\n\n\n\n\n")
    wb = openpyxl.Workbook()
    print(wb)
    wb_sheet = wb.active
    print("\n\n\n\n\n\nSheet is  created\n",wb_sheet,"\n\n\n\n\n\n\n\n")

    GD=frappe.get_doc("Global Defaults")
    company_name=GD.default_company
    address=frappe.get_doc("Address",GD.default_company+"-billing")
    company_address=address.address_line1+", "+address.address_line2+", "+address.city+", "+address.state+", "+address.county
    
    # Add Sheet Head for company Information
    """ Here we Merge cell A1 to I2 and show company inforamtion"""
    wb_sheet.merge_cells('A1:I2')
    cell=wb_sheet.cell(row=1,column=1)
    cell.value = company_name+"\n"+company_address
    cell.alignment = Alignment(horizontal='center', vertical='center')



    # Add column headers
    wb_sheet["A3"]="Quotation To"
    wb_sheet["B3"]="Customer Name"
    wb_sheet["C3"]="Transaction Date"
    wb_sheet["D3"]="Shift"
    wb_sheet["E3"]="Valid Till"
    wb_sheet["F3"]="Contact"
    wb_sheet["G3"]="Total"
    wb_sheet["H3"]="Tax"
    wb_sheet["I3"]="Grand Total"

    # Style Column head
    # style_head=wb_sheet.cell(row=1,column=9)
    # style_head.font=style_head.font.copy(bold=True)


    # Add data in workbook
    row = 4
    for data in datas:
        wb_sheet["A"+str(row)]=data.get("quotation_to")
        wb_sheet["B"+str(row)]=data.get("customer_name")
        wb_sheet["C"+str(row)]=data.get("transaction_date").strftime("%d-%b-%Y")
        wb_sheet["D"+str(row)]=data.get("shift")
        wb_sheet["E"+str(row)]=data.get("valid_till").strftime("%d-%b-%Y")
        wb_sheet["F"+str(row)]=data.get("contact_mobile")
        wb_sheet["G"+str(row)]=data.get("total")
        wb_sheet["H"+str(row)]=data.get("total_taxes_and_charges")
        wb_sheet["I"+str(row)]=data.get("grand_total")
        row+=1

    # Save workbook
    wb.save("/home/ruturaj/Downloads/report.xlsx")

    return
